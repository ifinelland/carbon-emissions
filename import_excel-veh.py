# import_excel_vehicles_concise.py
import os, math
import pandas as pd
from openpyxl import load_workbook
import mysql.connector

# ---------- CONFIG ----------
file_path = "Data/Zamboanga Branch/2025_Zam.xlsm"
sheet_name = "2.1b Fuel - Vehicles"
table_name = "Fuel_Veh"
category_name = "Fuel - Vehicles"

DB = {
    "host": "localhost",
    "user": "root",
    "password": "Nor@eb@ng99",
    "database": "carbon_emissions"
}
# ----------------------------

MONTH_MAP = {m[:3].lower(): i for i, m in enumerate(
    ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"], 1)}

def clean(val):
    if val is None or (isinstance(val, float) and math.isnan(val)):
        return None
    if isinstance(val, str):
        v = val.strip()
        return None if v.lower() in ("", "nan", "none") else v
    return val

def parse_month(raw):
    if isinstance(raw, (int, float)):
        m = int(raw)
        return m if 1 <= m <= 12 else None
    s = str(raw).strip().lower()[:3]
    return MONTH_MAP.get(s)

def get_pk(cursor, table):
    cursor.execute("SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS "
                   "WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME=%s AND COLUMN_KEY='PRI' LIMIT 1", (table,))
    r = cursor.fetchone()
    return r[0] if r else None

def get_or_create_tp(cursor, month, quarter, year):
    mnum = parse_month(month)
    label = f"{year}-{mnum:02d}" if mnum else None
    pk = get_pk(cursor, 'time_periods') or 'time_period_id'

    if label:
        cursor.execute(f"SELECT {pk} FROM time_periods WHERE label=%s", (label,))
        r = cursor.fetchone()
        if r: return r[0]

    month_name = None
    if mnum:
        import calendar
        month_name = calendar.month_name[mnum]
        cursor.execute(f"SELECT {pk} FROM time_periods WHERE year=%s AND month=%s", (year, month_name))
        r = cursor.fetchone()
        if r: return r[0]

    if quarter:
        q = str(quarter).strip().upper()
        cursor.execute(f"SELECT {pk} FROM time_periods WHERE year=%s AND quarter=%s LIMIT 1", (year, q))
        r = cursor.fetchone()
        if r: return r[0]

    cursor.execute("INSERT INTO time_periods (year, quarter, month, label) VALUES (%s,%s,%s,%s)",
                   (year, clean(quarter), month_name, label))
    return cursor.lastrowid

# --- Read Excel ---
wb = load_workbook(file_path, data_only=True)
if sheet_name not in wb.sheetnames: raise SystemExit(f"Sheet '{sheet_name}' not found")
ws = wb[sheet_name]
if table_name not in ws.tables: raise SystemExit(f"Table '{table_name}' not found. Available: {list(ws.tables.keys())}")

rows = [[c.value for c in r] for r in ws[ws.tables[table_name].ref]]
headers = rows[0]
values = [r for r in rows[1:] if any(c not in (None,"") for c in r)]
df = pd.DataFrame(values, columns=headers)

cols = {c.lower().replace(" ","_"): c for c in df.columns}
def fc(*names): return next((cols[n.lower().replace(" ","_")] for n in names if n.lower().replace(" ","_") in cols), None)

req_cols = [
    fc("Vehicle Type","Vehicle","vehicle_type"),
    fc("Fuel Type","Fuel","fuel_type"),
    fc("Month","month"),
    fc("Quarter","quarter","qtr"),
    fc("Consumption","amount","value"),
    fc("Unit","unit"),
    fc("Total Kilometers Travelled","Total kilometers travelled (all vehicles)","KM Travelled","total_kilometers_travelled"),
    fc("Unit2","unit2")
]

if None in req_cols: raise SystemExit(f"Missing required columns. Detected: {list(df.columns)}")

df = df[req_cols]
df.columns = ["vehicle_type","fuel_type","month","quarter","consumption","unit","total_kilometers_travelled","unit2"]
df["facility_type"] = "Office"
df = df[["facility_type"]+list(df.columns[:-1])]
df = df.applymap(clean)

# --- DB ---
conn = mysql.connector.connect(**DB)
cur = conn.cursor()
offices_pk = get_pk(cur,'offices') or 'office_id'
cats_pk = get_pk(cur,'categories') or 'category_id'
tp_pk = get_pk(cur,'time_periods') or 'time_period_id'

# --- Create table if not exists ---
cur.execute("""
CREATE TABLE IF NOT EXISTS fuel_vehicles (
    id INT AUTO_INCREMENT PRIMARY KEY,
    office_id INT,
    time_period_id INT,
    category_id INT,
    facility_type VARCHAR(100),
    vehicle_type VARCHAR(100),
    fuel_type VARCHAR(100),
    month VARCHAR(20),
    quarter VARCHAR(10),
    consumption DECIMAL(12,2),
    unit VARCHAR(20),
    total_kilometers_travelled DECIMAL(12,2),
    unit2 VARCHAR(20)
);""")

# --- Category ---
cur.execute(f"SELECT {cats_pk} FROM categories WHERE category_name=%s",(category_name,))
r=cur.fetchone()
if r: category_id=r[0]
else:
    cur.execute("INSERT INTO categories (category_name,description) VALUES (%s,%s)", (category_name,"Fuel consumption for vehicles"))
    category_id = cur.lastrowid
    conn.commit()

office_name = input("Enter office name: ").strip()
cur.execute(f"SELECT {offices_pk} FROM offices WHERE office_name=%s",(office_name,))
r=cur.fetchone()
if not r: raise SystemExit(f"Office '{office_name}' not found")
office_id = r[0]
year = int(input("Enter reporting year: "))

insert_sql = """INSERT INTO fuel_vehicles
(office_id,time_period_id,category_id,facility_type,vehicle_type,fuel_type,month,quarter,consumption,unit,total_kilometers_travelled,unit2)
VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"""

for _, row in df.iterrows():
    tp_id = get_or_create_tp(cur,row["month"],row["quarter"],year)
    cons = row["consumption"]
    kms = row["total_kilometers_travelled"]
    params = (office_id,tp_id,category_id,row["facility_type"],row["vehicle_type"],row["fuel_type"],
              row["month"],row["quarter"],cons,row["unit"],kms,row["unit2"])
    cur.execute(insert_sql, params)

conn.commit()
cur.close()
conn.close()
print(f"🎉 Done. Inserted {len(df)} rows into fuel_vehicles (office: {office_name}, year: {year}).")
