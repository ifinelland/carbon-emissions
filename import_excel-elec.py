# import_excel_electricity.py
import os, math
import pandas as pd
from openpyxl import load_workbook
import mysql.connector
import calendar

# ---------- CONFIG ----------
file_path = "Data/Head Office/2025_HO-FMED.xlsm"
sheet_name = "2.3a Electricity"
table_name = "Electricity"
category_name = "Electricity"

DB = {
    "host": "localhost",
    "user": "root",
    "password": "Nor@eb@ng99",
    "database": "carbon_emissions"
}
# ----------------------------

MONTH_MAP = {m[:3].lower(): i for i, m in enumerate(
    ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"], 1)}

def clean(val):
    if val is None or (isinstance(val, float) and math.isnan(val)):
        return None
    if isinstance(val, str):
        v = val.strip()
        return None if v.lower() in ("", "nan", "none") else v
    return val

def parse_month(raw):
    if isinstance(raw, (int, float)):
        m = int(raw)
        return m if 1 <= m <= 12 else None
    s = str(raw).strip().lower()[:3]
    return MONTH_MAP.get(s)

def get_pk(cursor, table):
    cursor.execute("SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS "
                   "WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME=%s AND COLUMN_KEY='PRI' LIMIT 1", (table,))
    r = cursor.fetchone()
    return r[0] if r else None

def get_or_create_tp(cursor, month, quarter, year):
    mnum = parse_month(month)
    label = f"{year}-{mnum:02d}" if mnum else None
    pk = get_pk(cursor, 'time_periods') or 'time_period_id'

    if label:
        cursor.execute(f"SELECT {pk} FROM time_periods WHERE label=%s", (label,))
        r = cursor.fetchone()
        if r: return r[0]

    month_name = None
    if mnum:
        month_name = calendar.month_name[mnum]
        cursor.execute(f"SELECT {pk} FROM time_periods WHERE year=%s AND month=%s", (year, month_name))
        r = cursor.fetchone()
        if r: return r[0]

    if quarter:
        q = str(quarter).strip().upper()
        cursor.execute(f"SELECT {pk} FROM time_periods WHERE year=%s AND quarter=%s LIMIT 1", (year, q))
        r = cursor.fetchone()
        if r: return r[0]

    cursor.execute("INSERT INTO time_periods (year, quarter, month, label) VALUES (%s,%s,%s,%s)",
                   (year, clean(quarter), month_name, label))
    return cursor.lastrowid

# --- Read Excel ---
wb = load_workbook(file_path, data_only=True)
if sheet_name not in wb.sheetnames: 
    raise SystemExit(f"Sheet '{sheet_name}' not found")
ws = wb[sheet_name]
if table_name not in ws.tables: 
    raise SystemExit(f"Table '{table_name}' not found. Available: {list(ws.tables.keys())}")

rows = [[c.value for c in r] for r in ws[ws.tables[table_name].ref]]
headers = rows[0]
values = [r for r in rows[1:] if any(c not in (None,"") for c in r)]
df = pd.DataFrame(values, columns=headers)

# column lookup helper
cols = {c.lower().replace(" ","_"): c for c in df.columns}
def fc(*names): 
    return next((cols[n.lower().replace(" ","_")] for n in names if n.lower().replace(" ","_") in cols), None)

# check if Facility Type exists
facility_col = fc("Facility Type","facility_type")

req_cols = [
    fc("Month","month"),
    fc("Quarter","quarter","qtr"),
    fc("Consumption (kWh)","consumption","kwh"),
    fc("Cost (PHP)","cost","php"),
    fc("Utility Provider","provider","utility_provider")
]

if None in req_cols: 
    missing = [n for n, c in zip(["month","quarter","consumption_kwh","cost_php","utility_provider"], req_cols) if c is None]
    raise SystemExit(f"Missing required columns: {missing}. Detected: {list(df.columns)}")

# select required cols
df = df[[facility_col] + req_cols] if facility_col else df[req_cols]

# rename columns
new_cols = (["facility_type","month","quarter","consumption_kwh","cost_php","utility_provider"]
            if facility_col else ["month","quarter","consumption_kwh","cost_php","utility_provider"])
df.columns = new_cols

# add default facility_type if missing
if "facility_type" not in df.columns:
    df.insert(0, "facility_type", "Office")

df = df.applymap(clean)

# --- DB ---
conn = mysql.connector.connect(**DB)
cur = conn.cursor()
offices_pk = get_pk(cur,'offices') or 'office_id'
cats_pk = get_pk(cur,'categories') or 'category_id'
tp_pk = get_pk(cur,'time_periods') or 'time_period_id'

# --- Create table if not exists ---
cur.execute("""
CREATE TABLE IF NOT EXISTS electricity (
    id INT AUTO_INCREMENT PRIMARY KEY,
    office_id INT,
    time_period_id INT,
    category_id INT,
    facility_type VARCHAR(100),
    month VARCHAR(20),
    quarter VARCHAR(10),
    consumption_kwh DECIMAL(12,2),
    cost_php DECIMAL(12,2),
    utility_provider VARCHAR(100)
);""")

# --- Category ---
cur.execute(f"SELECT {cats_pk} FROM categories WHERE category_name=%s",(category_name,))
r=cur.fetchone()
if r: 
    category_id=r[0]
else:
    cur.execute("INSERT INTO categories (category_name,description) VALUES (%s,%s)", 
                (category_name,"Electricity consumption"))
    category_id = cur.lastrowid
    conn.commit()

office_name = input("Enter office name: ").strip()
cur.execute(f"SELECT {offices_pk} FROM offices WHERE office_name=%s",(office_name,))
r=cur.fetchone()
if not r: raise SystemExit(f"Office '{office_name}' not found")
office_id = r[0]
year = int(input("Enter reporting year: "))

insert_sql = """INSERT INTO electricity
(office_id,time_period_id,category_id,facility_type,month,quarter,consumption_kwh,cost_php,utility_provider)
VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)"""

params_list = []
for _, row in df.iterrows():
    tp_id = get_or_create_tp(cur,row["month"],row["quarter"],year)
    params = (office_id,tp_id,category_id,row["facility_type"],row["month"],row["quarter"],
              row["consumption_kwh"],row["cost_php"],row["utility_provider"])
    params_list.append(params)

cur.executemany(insert_sql, params_list)

conn.commit()
cur.close()
conn.close()
print(f"🎉 Done. Inserted {len(df)} rows into electricity (office: {office_name}, year: {year}).")
