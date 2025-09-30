# import_fuel_buildings.py
import os
from openpyxl import load_workbook
import pandas as pd
import mysql.connector

# ---------- CONFIG - edit these ----------
file_path = "/Users/finellajianna/carbon-emissions/Data/Zamboanga Branch/2024_Zam.xlsm"
sheet_name = "2.1a Fuel - Buildings"
table_name = "Fuel_Blg"
reporting_year = 2024

DB = {
    "host": "localhost",
    "user": "root",
    "password": "Nor@eb@ng99",
    "database": "carbon_emissions"
}
# -----------------------------------------

# --- helpers ---
def normalize_colname(s):
    return "" if s is None else s.strip().lower().replace(".", "").replace(" ", "_")

MONTH_MAP = {
    "jan":1, "january":1,
    "feb":2, "february":2,
    "mar":3, "march":3,
    "apr":4, "april":4,
    "may":5,
    "jun":6, "june":6,
    "jul":7, "july":7,
    "aug":8, "august":8,
    "sep":9, "sept":9, "september":9,
    "oct":10, "october":10,
    "nov":11, "november":11,
    "dec":12, "december":12
}

def parse_month_to_number(raw):
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        try:
            m = int(raw)
            if 1 <= m <= 12: return m
        except: 
            pass
    s = str(raw).strip().lower()
    s_clean = s.replace(".", "")
    if s_clean.isdigit():
        m = int(s_clean)
        if 1 <= m <= 12:
            return m
    if s_clean[:3] in MONTH_MAP:
        return MONTH_MAP.get(s_clean[:3])
    return MONTH_MAP.get(s_clean)

def get_pk_column(cursor, table_name):
    cursor.execute("""
        SELECT COLUMN_NAME
        FROM INFORMATION_SCHEMA.COLUMNS
        WHERE TABLE_SCHEMA = DATABASE()
          AND TABLE_NAME = %s
          AND COLUMN_KEY = 'PRI'
        LIMIT 1
    """, (table_name,))
    r = cursor.fetchone()
    return r[0] if r else None

def get_or_create_time_period(cursor, month_raw, quarter_raw, year=reporting_year):
    mnum = parse_month_to_number(month_raw)
    label = f"{year}-{mnum:02d}" if mnum else None
    pk = get_pk_column(cursor, 'time_periods') or 'time_period_id'

    if label:
        cursor.execute(f"SELECT {pk} FROM time_periods WHERE label = %s", (label,))
        r = cursor.fetchone()
        if r:
            return r[0]

    month_full = None
    if mnum:
        import calendar
        month_full = calendar.month_name[mnum]
        cursor.execute(f"SELECT {pk} FROM time_periods WHERE year = %s AND month = %s", (year, month_full))
        r = cursor.fetchone()
        if r:
            return r[0]

    if quarter_raw:
        qnorm = str(quarter_raw).strip().upper()
        cursor.execute(f"SELECT {pk} FROM time_periods WHERE year = %s AND quarter = %s LIMIT 1", (year, qnorm))
        r = cursor.fetchone()
        if r:
            return r[0]

    month_val = month_full if month_full else (str(month_raw).strip() if month_raw is not None else None)
    qval = str(quarter_raw).strip() if quarter_raw is not None else None
    cursor.execute("INSERT INTO time_periods (year, quarter, month, label) VALUES (%s,%s,%s,%s)",
                   (year, qval, month_val, label))
    return cursor.lastrowid

# --- 1. Read Excel table using openpyxl ---
wb = load_workbook(file_path, data_only=True)
if sheet_name not in wb.sheetnames:
    raise SystemExit(f"Sheet '{sheet_name}' not found in {file_path}")
ws = wb[sheet_name]

if table_name not in ws.tables:
    raise SystemExit(f"Table '{table_name}' not found in sheet '{sheet_name}'. Available tables: {list(ws.tables.keys())}")

tbl = ws.tables[table_name]
ref = tbl.ref
cells = ws[ref]
rows = [[cell.value for cell in row] for row in cells]
if not rows or len(rows) < 2:
    raise SystemExit("No data found in the named table.")

headers = rows[0]
values = rows[1:]
df = pd.DataFrame(values, columns=headers)

# normalize column lookup
cols_norm = {normalize_colname(c): c for c in df.columns}
def find_col(*candidates):
    for cand in candidates:
        nc = normalize_colname(cand)
        if nc in cols_norm:
            return cols_norm[nc]
    return None

# --- column detection (facility_type optional) ---
fuel_col     = find_col("Fuel Type", "Fuel", "fuel_type")
month_col    = find_col("Month", "month")
quarter_col  = find_col("Quarter", "quarter", "qtr")
cons_col     = find_col("Consumption", "consumption", "amount", "value")
unit_col     = find_col("Unit", "unit")
facility_col = find_col("Facility Type", "Facility", "facility_type")

if not (fuel_col and month_col and quarter_col and cons_col and unit_col):
    print("Detected columns:", df.columns.tolist())
    raise SystemExit("Could not find the required columns in the named table. Check header names.")

if facility_col:
    df = df[[facility_col, fuel_col, month_col, quarter_col, cons_col, unit_col]]
    df.columns = ["facility_type", "fuel_type", "month", "quarter", "consumption", "unit"]
else:
    df = df[[fuel_col, month_col, quarter_col, cons_col, unit_col]]
    df.columns = ["fuel_type", "month", "quarter", "consumption", "unit"]
    df["facility_type"] = "Office"
    df = df[["facility_type", "fuel_type", "month", "quarter", "consumption", "unit"]]

df = df.where(pd.notnull(df), None)

print("✅ Extracted DataFrame (first 5 rows):")
print(df.head())

# --- 2. Connect to DB ---
conn = mysql.connector.connect(**DB)
cursor = conn.cursor()

offices_pk = get_pk_column(cursor, 'offices') or 'office_id'
cats_pk    = get_pk_column(cursor, 'categories') or 'category_id'
tp_pk      = get_pk_column(cursor, 'time_periods') or 'time_period_id'
print("Detected PK columns:", {"offices": offices_pk, "categories": cats_pk, "time_periods": tp_pk})

# --- 3. Create fuel_buildings table if not exists ---
cursor.execute("""
CREATE TABLE IF NOT EXISTS fuel_buildings (
    id INT AUTO_INCREMENT PRIMARY KEY,
    office_id INT,
    time_period_id INT,
    category_id INT,
    facility_type VARCHAR(100),
    fuel_type VARCHAR(100),
    month VARCHAR(20),
    quarter VARCHAR(10),
    consumption DECIMAL(12,2),
    unit VARCHAR(20)
);
""")


# --- 4. Resolve category_id ---
cursor.execute(f"SELECT {cats_pk} FROM categories WHERE category_name = %s", ("Fuel - Buildings",))
r = cursor.fetchone()
if r:
    category_id = r[0]
else:
    cursor.execute("INSERT INTO categories (category_name, description) VALUES (%s,%s)",
                   ("Fuel - Buildings", "Fuel consumption for buildings/stationary sources"))
    category_id = cursor.lastrowid
    conn.commit()
print("Using category_id =", category_id)

# --- 5. Resolve office_id ---
office_name = input("Enter the office name (must exist in offices): ").strip()
cursor.execute(f"SELECT {offices_pk} FROM offices WHERE office_name = %s", (office_name,))
r = cursor.fetchone()
if not r:
    raise SystemExit(f"Office '{office_name}' not found in offices table. Please add it first or use correct name.")
office_id = r[0]
print("Resolved office_id =", office_id)

# --- 5b. Reporting year ---
year = int(input("Enter reporting year (e.g. 2024): "))

# --- 6. Insert rows ---
insert_sql = """
INSERT INTO fuel_buildings
(office_id, time_period_id, category_id, facility_type, fuel_type, month, quarter, consumption, unit)
VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
"""
inserted = 0
for _, row in df.iterrows():
    month_raw = row["month"]
    quarter_raw = row["quarter"]
    tp_id = get_or_create_time_period(cursor, month_raw, quarter_raw, year=year)
    cons = None
    try:
        cons = None if row["consumption"] is None else float(row["consumption"])
    except:
        cons = None
    params = (office_id, tp_id, category_id,
              row["facility_type"], row["fuel_type"], row["month"], row["quarter"],
              cons, row["unit"])
    cursor.execute(insert_sql, params)
    inserted += 1

conn.commit()
cursor.close()
conn.close()

print(f"🎉 Done. Inserted {inserted} rows into fuel_buildings (office: {office_name}, year: {year}).")
